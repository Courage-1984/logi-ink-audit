"""Historical trend analysis against previous Excel audit workbooks."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

REPORT_NAME_PATTERN: re.Pattern[str] = re.compile(
    r"^live-audit-(\d{2}-\d{2}-\d{4})\.xlsx$",
    re.IGNORECASE,
)

BASELINE_STATUS: str = "Initial baseline audit — no previous trend data available"


def get_latest_previous_report(
    output_dir: str | Path,
    *,
    as_of: datetime | None = None,
) -> Path | None:
    """
    Locate the most recent `live-audit-*.xlsx` excluding today's run.

    Returns None when no prior workbook exists (graceful baseline fallback).
    """
    directory = Path(output_dir)
    if not directory.is_dir():
        return None

    as_of = as_of or datetime.now()
    today_stamp = as_of.strftime("%d-%m-%Y")
    candidates: list[tuple[datetime, Path]] = []

    for path in directory.glob("live-audit-*.xlsx"):
        match = REPORT_NAME_PATTERN.match(path.name)
        if not match:
            continue
        stamp = match.group(1)
        if stamp == today_stamp:
            continue
        try:
            parsed = datetime.strptime(stamp, "%d-%m-%Y")
        except ValueError:
            continue
        candidates.append((parsed, path))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (item[0], item[1].stat().st_mtime), reverse=True)
    latest = candidates[0][1]
    logger.info("Previous audit workbook located: %s", latest.name)
    return latest


def extract_critical_issues_from_workbook(workbook_path: Path) -> list[str]:
    """Reconstruct critical-failure lines from a prior Detailed Results sheet."""
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — graceful historical fallback
        logger.warning("Unable to read previous workbook %s: %s", workbook_path, exc)
        return []

    if "Detailed Results" not in workbook.sheetnames:
        workbook.close()
        return []

    worksheet = workbook["Detailed Results"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        workbook.close()
        return []

    header_map = {
        str(name).strip(): index
        for index, name in enumerate(headers)
        if name is not None
    }

    def cell(row: tuple[Any, ...], header: str, default: Any = "") -> Any:
        index = header_map.get(header)
        if index is None or index >= len(row):
            return default
        value = row[index]
        return default if value is None else value

    def as_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value).strip().upper()
        return text in {"TRUE", "1", "YES", "PASS"}

    lines: list[str] = []
    for row in rows:
        url = str(cell(row, "URL", "")).strip()
        if not url:
            continue

        try:
            status = int(cell(row, "Status", 0) or 0)
        except (TypeError, ValueError):
            status = 0
        if status >= 400 or status == 0:
            label = status if status else "no response"
            lines.append(f"{url} - HTTP {label}")

        if not as_bool(cell(row, "Title Present", True)):
            lines.append(f"{url} - Missing <title> tag")

        if not as_bool(cell(row, "Meta Desc Present", True)):
            lines.append(f"{url} - Missing meta description")

        try:
            alt_count = int(cell(row, "Missing Alt Text Count", 0) or 0)
        except (TypeError, ValueError):
            alt_count = 0
        if alt_count > 0:
            lines.append(f"{url} - Missing alt text on images ({alt_count})")

        schema_valid = as_bool(cell(row, "Schema Valid", False))
        parse_fail = as_bool(cell(row, "JSON-LD Parse Fail", False))
        schemas = str(cell(row, "Detected Schema Types", "")).strip()
        if parse_fail:
            lines.append(f"{url} - Malformed JSON-LD structured data (Fail)")
        elif not schema_valid or not schemas:
            lines.append(f"{url} - Missing structured data (JSON-LD)")

        has_org = as_bool(cell(row, "Has Organization", False))
        has_local = as_bool(cell(row, "Has LocalBusiness", False))
        if not has_org and not has_local:
            lines.append(f"{url} - Missing LocalBusiness / Organization schema")

        if not as_bool(cell(row, "HTML Lang en-ZA", True)):
            lang = str(cell(row, "HTML Lang", "(missing)") or "(missing)")
            lines.append(f"{url} - Failed en-ZA localisation (html lang={lang})")

        currency = str(cell(row, "priceCurrency", "") or "").strip()
        local_status = str(cell(row, "Local SEO Compliance (Pass/Fail)", "") or "").strip()
        if local_status.lower() == "fail":
            if currency and currency.upper() != "ZAR":
                lines.append(
                    f"{url} - Failed ZAR localisation (priceCurrency={currency})"
                )
            elif not currency:
                # Soft signal — only flag ZAR when Service schemas imply currency needed
                schemas_lower = schemas.lower()
                if "service" in schemas_lower or "product" in schemas_lower or "offer" in schemas_lower:
                    lines.append(
                        f"{url} - Failed ZAR localisation (priceCurrency=(missing))"
                    )

            region = str(cell(row, "addressRegion", "") or "")
            locality = str(cell(row, "addressLocality", "") or "")
            country = str(cell(row, "addressCountry", "") or "")
            if (
                "gauteng" not in region.lower()
                or "pretoria" not in locality.lower()
                or country.upper() not in {"ZA", "ZAF"}
            ):
                lines.append(
                    f"{url} - Failed South African localisation "
                    f"(expected Pretoria / Gauteng / ZA)"
                )

    workbook.close()

    seen: set[str] = set()
    unique: list[str] = []
    for line in lines:
        if line not in seen:
            seen.add(line)
            unique.append(line)
    return unique


def compute_delta(
    previous_file: Path,
    current_seo_results: list,
    current_geo_results: list,
    current_tech_results: list,
) -> dict[str, Any]:
    """
    Compare current critical failures against a previous workbook.

    Returns resolved / new / persistent issue lists plus a human-readable summary.
    """
    previous_issues = set(extract_critical_issues_from_workbook(previous_file))
    from .reporter import summarise_critical_failures

    current_issues = set(
        summarise_critical_failures(
            current_seo_results,
            current_geo_results,
            current_tech_results,
        )
    )

    resolved = sorted(previous_issues - current_issues)
    new_issues = sorted(current_issues - previous_issues)
    persistent = sorted(previous_issues & current_issues)

    summary = (
        f"+{len(new_issues)} New Issues, "
        f"-{len(resolved)} Resolved Issues, "
        f"{len(persistent)} Persistent Issues"
    )

    return {
        "has_previous": True,
        "previous_file": str(previous_file),
        "previous_name": previous_file.name,
        "status": summary,
        "summary_line": summary,
        "resolved_count": len(resolved),
        "new_count": len(new_issues),
        "persistent_count": len(persistent),
        "resolved_issues": resolved,
        "new_issues": new_issues,
        "persistent_issues": persistent,
        "previous_issue_count": len(previous_issues),
        "current_issue_count": len(current_issues),
    }


def build_baseline_delta() -> dict[str, Any]:
    """Neutral delta payload when no previous report exists."""
    return {
        "has_previous": False,
        "previous_file": "",
        "previous_name": "",
        "status": BASELINE_STATUS,
        "summary_line": BASELINE_STATUS,
        "resolved_count": 0,
        "new_count": 0,
        "persistent_count": 0,
        "resolved_issues": [],
        "new_issues": [],
        "persistent_issues": [],
        "previous_issue_count": 0,
        "current_issue_count": 0,
    }


def analyse_historical_trends(
    output_dir: str | Path,
    current_seo_results: Sequence[Any],
    current_geo_results: Sequence[Any],
    current_tech_results: Sequence[Any],
    *,
    as_of: datetime | None = None,
) -> dict[str, Any]:
    """Orchestrate previous-report lookup and delta computation."""
    from .reporter import summarise_critical_failures

    logger.info("Analysing historical trends...")
    previous = get_latest_previous_report(output_dir, as_of=as_of)
    if previous is None:
        logger.info(BASELINE_STATUS)
        delta = build_baseline_delta()
        delta["current_issue_count"] = len(
            summarise_critical_failures(
                list(current_seo_results),
                list(current_geo_results),
                list(current_tech_results),
            )
        )
        return delta

    delta = compute_delta(
        previous,
        list(current_seo_results),
        list(current_geo_results),
        list(current_tech_results),
    )
    logger.info("Delta summary: %s", delta["summary_line"])
    return delta
