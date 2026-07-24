"""Aggregate audit results into an unbranded Excel workbook with formula summaries."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .auditor_geo_aeo import GeoAeoAuditResult
from .auditor_seo import SeoAuditResult
from .auditor_technical import TechnicalAuditResult

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SECTION_FILL = PatternFill("solid", fgColor="374151")
GOOD_FILL = PatternFill("solid", fgColor="D1FAE5")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
BAD_FILL = PatternFill("solid", fgColor="FECACA")
TRUE_FONT = Font(color="065F46")
FALSE_FONT = Font(color="991B1B")
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

DETAIL_COLUMNS: list[tuple[str, str]] = [
    ("url", "URL"),
    ("status_code", "Status"),
    ("latency_ms", "Latency (ms)"),
    ("redirected", "Redirected"),
    ("html_lang", "HTML Lang"),
    ("html_lang_en_za", "HTML Lang en-ZA"),
    ("title", "Title"),
    ("title_length", "Title Length"),
    ("title_present", "Title Present"),
    ("title_length_ok", "Title Length OK"),
    ("title_brand_ok", "Title Brand OK"),
    ("meta_description_present", "Meta Desc Present"),
    ("meta_description_length", "Meta Desc Length"),
    ("meta_description_length_ok", "Meta Desc Length OK"),
    ("terminology_ok", "Terminology OK"),
    ("terminology_issues", "Terminology Issues"),
    ("h1_count", "H1 Count"),
    ("heading_hierarchy_ok", "Heading Hierarchy OK"),
    ("heading_skip_detected", "Heading Skip"),
    ("missing_alt_count", "Missing Alt Text Count"),
    ("empty_alt_count", "Empty Alt Count"),
    ("alt_coverage_ok", "Alt Coverage OK"),
    ("internal_link_count", "Internal Link Count"),
    ("external_link_count", "External Link Count"),
    ("internal_link_ratio", "Internal Link Ratio"),
    ("link_density_ok", "Link Density OK"),
    ("orphan_risk", "Orphan Risk"),
    ("landmark_coverage_ok", "Landmarks (main/nav/header)"),
    ("payload_kb", "Payload (KB)"),
    ("script_count", "Script Count"),
    ("css_count", "CSS Count"),
    ("dom_node_count", "DOM Node Count"),
    ("dom_weight_status", "DOM Weight Status"),
    ("dom_weight_ok", "DOM Weight OK"),
    ("og_title_present", "OG Title"),
    ("og_description_present", "OG Description"),
    ("og_image_present", "OG Image"),
    ("og_image_path_ok", "OG Image Path OK"),
    ("og_image", "OG Image URL"),
    ("canonical_present", "Canonical Present"),
    ("canonical_ok", "Canonical OK"),
    ("seo_score", "SEO Score"),
    ("technical_score", "Technical Score"),
    ("seo_issues", "SEO Issues"),
    ("technical_issues", "Technical Issues"),
    ("detected_schema_types", "Detected Schema Types"),
    ("schema_valid", "Schema Valid"),
    ("json_ld_parse_fail", "JSON-LD Parse Fail"),
    ("has_organization", "Has Organization"),
    ("has_local_business", "Has LocalBusiness"),
    ("has_service", "Has Service"),
    ("has_website", "Has WebSite"),
    ("has_breadcrumb", "Has Breadcrumb"),
    ("address_region", "addressRegion"),
    ("address_locality", "addressLocality"),
    ("address_country", "addressCountry"),
    ("price_currency", "priceCurrency"),
    ("local_seo_status", "Local SEO Compliance (Pass/Fail)"),
    ("local_seo_compliance", "Local SEO Compliant"),
    ("localisation_gaps", "Localisation Gaps"),
    ("semantic_tags", "Semantic Tags"),
    ("semantic_coverage_ok", "Semantic Coverage OK"),
    ("has_definition_list", "Definition List (dl/dt/dd)"),
    ("has_details_summary", "Details/Summary Accordion"),
    ("has_faq_schema", "FAQ Schema"),
    ("has_question_headings", "Question Headings"),
    ("question_heading_count", "Question Heading Count"),
    ("ai_chunking_ready", "AI Chunking Ready"),
    ("direct_answer_ready", "Direct-Answer Ready"),
    ("has_definition_pattern", "Definition Pattern"),
    ("has_structured_lists", "Structured Lists"),
    ("has_summary_block", "Summary Block"),
    ("nap_complete", "NAP Complete"),
    ("has_same_as", "Has sameAs"),
    ("entity_mapping_ok", "Entity Mapping OK"),
    ("geo_score", "GEO Score"),
    ("aeo_score", "AEO Score"),
    ("geo_issues", "GEO Gaps"),
    ("aeo_issues", "AEO Gaps"),
]


def summarise_critical_failures(
    seo_results: Sequence[SeoAuditResult],
    geo_results: Sequence[GeoAeoAuditResult],
    technical_results: Sequence[TechnicalAuditResult] | None = None,
) -> list[str]:
    """
    Compile critical SEO / GEO failure lines for email and console summaries.

    Critical failures include HTTP errors, missing titles / meta descriptions,
    missing image alt text, absent or invalid JSON-LD, and failed South African
    localisation (en-ZA / ZAR / Pretoria–Gauteng–ZA address signals).
    """
    technical_by_url = {
        result.url: result for result in (technical_results or [])
    }
    geo_by_url = {result.url: result for result in geo_results}
    lines: list[str] = []

    for seo in seo_results:
        url = seo.url
        geo = geo_by_url.get(url)
        technical = technical_by_url.get(url)

        if seo.status_code >= 400 or seo.status_code == 0:
            label = seo.status_code if seo.status_code else "no response"
            lines.append(f"{url} - HTTP {label}")

        if not seo.title_present:
            lines.append(f"{url} - Missing <title> tag")

        if not seo.meta_description_present:
            lines.append(f"{url} - Missing meta description")

        alt_issues = 0
        if technical is not None:
            alt_issues = technical.alt_issues_count
        elif seo.missing_alt_count:
            alt_issues = seo.missing_alt_count
        if alt_issues > 0:
            lines.append(f"{url} - Missing alt text on images ({alt_issues})")

        if geo is None:
            lines.append(f"{url} - Missing structured data (JSON-LD)")
            continue

        if geo.json_ld_blocks == 0 or not geo.schema_valid:
            if geo.json_ld_parse_fail:
                lines.append(f"{url} - Malformed JSON-LD structured data (Fail)")
            else:
                lines.append(f"{url} - Missing structured data (JSON-LD)")

        if not geo.has_local_business and not geo.has_organization:
            lines.append(f"{url} - Missing LocalBusiness / Organization schema")

        if not geo.html_lang_en_za:
            lang_display = geo.html_lang or "(missing)"
            lines.append(f"{url} - Failed en-ZA localisation (html lang={lang_display})")

        if geo.price_currency_applicable and not geo.price_currency_ok:
            currency_display = geo.price_currency or "(missing)"
            lines.append(f"{url} - Failed ZAR localisation (priceCurrency={currency_display})")

        if not geo.local_seo_compliance:
            # Surface address-level gaps when overall localisation failed
            if not geo.address_region_ok or not geo.address_locality_ok or not geo.address_country_ok:
                lines.append(
                    f"{url} - Failed South African localisation "
                    f"(expected Pretoria / Gauteng / ZA)"
                )
            elif geo.localisation_gaps:
                # Avoid duplicating en-ZA / ZAR lines already added above
                remaining = [
                    gap
                    for gap in geo.localisation_gaps
                    if "en-ZA" not in gap and "priceCurrency" not in gap and "html lang" not in gap
                ]
                for gap in remaining[:2]:
                    lines.append(f"{url} - {gap}")

    # De-duplicate whilst preserving order
    seen: set[str] = set()
    unique: list[str] = []
    for line in lines:
        if line not in seen:
            seen.add(line)
            unique.append(line)
    return unique


class AuditReporter:
    """Build a formula-driven Excel scorecard from SEO / GEO / AEO / technical results."""

    def __init__(self, output_dir: str | Path = "audit-reports") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        seo_results: Sequence[SeoAuditResult],
        geo_results: Sequence[GeoAeoAuditResult],
        technical_results: Sequence[TechnicalAuditResult] | None = None,
        *,
        base_url: str,
        report_date: datetime | None = None,
        delta: dict[str, Any] | None = None,
    ) -> Path:
        report_date = report_date or datetime.now()
        stamp = report_date.strftime("%d-%m-%Y")
        path = self.output_dir / f"live-audit-{stamp}.xlsx"

        merged = self._merge(seo_results, geo_results, technical_results or [])
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Overview"
        detail = workbook.create_sheet("Detailed Results")

        self._build_overview(overview, merged, base_url, report_date, delta=delta or {})
        self._build_detail(detail, merged)

        workbook.save(path)
        return path

    def summarise_critical_failures(
        self,
        seo_results: Sequence[SeoAuditResult],
        geo_results: Sequence[GeoAeoAuditResult],
        technical_results: Sequence[TechnicalAuditResult] | None = None,
    ) -> list[str]:
        """Instance wrapper around the module-level critical-failure summariser."""
        return summarise_critical_failures(seo_results, geo_results, technical_results)

    def _merge(
        self,
        seo_results: Sequence[SeoAuditResult],
        geo_results: Sequence[GeoAeoAuditResult],
        technical_results: Sequence[TechnicalAuditResult],
    ) -> list[dict[str, Any]]:
        geo_by_url = {result.url: result for result in geo_results}
        technical_by_url = {result.url: result for result in technical_results}
        rows: list[dict[str, Any]] = []
        for seo in seo_results:
            row = seo.to_dict()
            geo = geo_by_url.get(seo.url)
            technical = technical_by_url.get(seo.url)
            if geo:
                row.update(geo.to_dict())
            else:
                row.setdefault("schema_types", "")
                row.setdefault("detected_schema_types", "")
                row.setdefault("schema_valid", False)
                row.setdefault("geo_score", 0.0)
                row.setdefault("aeo_score", 0.0)
                row.setdefault("geo_issues", "")
                row.setdefault("aeo_issues", "")
                row.setdefault("local_seo_status", "Fail")
                row.setdefault("local_seo_compliance", False)
                row.setdefault("html_lang", "")
                row.setdefault("localisation_gaps", "")
            if technical:
                row.update(technical.to_dict())
                # Prefer technical auditor's stricter missing+empty alt count for the report column
                row["missing_alt_count"] = technical.alt_issues_count
            else:
                row.setdefault("empty_alt_count", 0)
                row.setdefault("internal_link_count", 0)
                row.setdefault("external_link_count", 0)
                row.setdefault("internal_link_ratio", 0.0)
                row.setdefault("link_density_ok", False)
                row.setdefault("orphan_risk", True)
                row.setdefault("landmark_coverage_ok", False)
                row.setdefault("technical_score", 0.0)
                row.setdefault("technical_issues", "")
                row.setdefault("alt_coverage_ok", False)
                row.setdefault("payload_kb", 0.0)
                row.setdefault("script_count", 0)
                row.setdefault("css_count", 0)
                row.setdefault("dom_node_count", 0)
                row.setdefault("dom_weight_status", "Unknown")
                row.setdefault("dom_weight_ok", False)
            row.setdefault("terminology_issues", "")
            row.setdefault("terminology_ok", True)
            row.setdefault("has_definition_list", False)
            row.setdefault("has_details_summary", False)
            row.setdefault("has_faq_schema", False)
            row.setdefault("has_question_headings", False)
            row.setdefault("question_heading_count", 0)
            row.setdefault("ai_chunking_ready", False)
            rows.append(row)
        return rows

    def _build_overview(
        self,
        worksheet: Worksheet,
        rows: list[dict[str, Any]],
        base_url: str,
        report_date: datetime,
        *,
        delta: dict[str, Any] | None = None,
    ) -> None:
        delta = delta or {
            "has_previous": False,
            "summary_line": "Initial baseline audit — no previous trend data available",
            "resolved_count": 0,
            "new_count": 0,
            "persistent_count": 0,
            "previous_name": "",
            "resolved_issues": [],
            "new_issues": [],
            "persistent_issues": [],
        }

        worksheet["A1"] = "Website Audit Overview"
        worksheet["A1"].font = Font(bold=True, size=16, color="111827", name="Calibri")
        worksheet.merge_cells("A1:D1")

        worksheet["A2"] = "Target site"
        worksheet["B2"] = base_url
        worksheet["A3"] = "Report date"
        worksheet["B3"] = report_date.strftime("%d-%m-%Y %H:%M")
        worksheet["A4"] = "Pages audited"
        worksheet["B4"] = "=COUNTA('Detailed Results'!A:A)-1"

        for label_cell in ("A2", "A3", "A4"):
            worksheet[label_cell].font = Font(bold=True)

        worksheet["A6"] = "Pillar"
        worksheet["B6"] = "Average compliance %"
        worksheet["C6"] = "Pass rate (score ≥ 70)"
        for column in ("A", "B", "C"):
            worksheet[f"{column}6"].fill = HEADER_FILL
            worksheet[f"{column}6"].font = HEADER_FONT

        seo_col = get_column_letter(self._detail_col_index("seo_score"))
        geo_col = get_column_letter(self._detail_col_index("geo_score"))
        aeo_col = get_column_letter(self._detail_col_index("aeo_score"))
        technical_col = get_column_letter(self._detail_col_index("technical_score"))
        status_col = get_column_letter(self._detail_col_index("status_code"))
        schema_col = get_column_letter(self._detail_col_index("schema_valid"))
        heading_col = get_column_letter(self._detail_col_index("heading_hierarchy_ok"))
        local_col = get_column_letter(self._detail_col_index("local_seo_status"))
        term_col = get_column_letter(self._detail_col_index("terminology_ok"))
        chunk_col = get_column_letter(self._detail_col_index("ai_chunking_ready"))
        weight_col = get_column_letter(self._detail_col_index("dom_weight_status"))
        last_data_hint = "1000"

        pillars = [
            ("SEO", seo_col),
            ("GEO", geo_col),
            ("AEO", aeo_col),
            ("Technical / A11y", technical_col),
        ]
        for index, (name, column) in enumerate(pillars, start=7):
            worksheet[f"A{index}"] = name
            worksheet[f"B{index}"] = (
                f"=IFERROR(AVERAGE('Detailed Results'!{column}2:{column}{last_data_hint}),0)"
            )
            worksheet[f"C{index}"] = (
                f"=IFERROR(COUNTIF('Detailed Results'!{column}2:{column}{last_data_hint},\">=70\")"
                f"/MAX(1,COUNTA('Detailed Results'!A2:A{last_data_hint})),0)"
            )
            worksheet[f"B{index}"].number_format = "0.0"
            worksheet[f"C{index}"].number_format = "0.0%"

        worksheet["A11"] = "Overall compliance %"
        worksheet["B11"] = "=IFERROR(AVERAGE(B7:B10),0)"
        worksheet["B11"].number_format = "0.0"
        worksheet["A11"].font = Font(bold=True)
        worksheet["B11"].font = Font(bold=True, size=14)

        worksheet["A13"] = "Operational flags"
        worksheet["A13"].fill = SECTION_FILL
        worksheet["A13"].font = Font(bold=True, color="FFFFFF")
        worksheet.merge_cells("A13:B13")

        worksheet["A14"] = "HTTP errors (status ≥ 400)"
        worksheet["B14"] = (
            f"=COUNTIF('Detailed Results'!{status_col}2:{status_col}{last_data_hint},\">=400\")"
        )
        worksheet["A15"] = "Pages missing valid schema"
        worksheet["B15"] = (
            f"=COUNTIF('Detailed Results'!{schema_col}2:{schema_col}{last_data_hint},FALSE)"
        )
        worksheet["A16"] = "Pages with heading hierarchy issues"
        worksheet["B16"] = (
            f"=COUNTIF('Detailed Results'!{heading_col}2:{heading_col}{last_data_hint},FALSE)"
        )
        worksheet["A17"] = "Pages failing SA localisation"
        worksheet["B17"] = (
            f"=COUNTIF('Detailed Results'!{local_col}2:{local_col}{last_data_hint},\"Fail\")"
        )
        worksheet["A18"] = "Pages with terminology issues"
        worksheet["B18"] = (
            f"=COUNTIF('Detailed Results'!{term_col}2:{term_col}{last_data_hint},FALSE)"
        )
        worksheet["A19"] = "Pages not AI-chunking ready"
        worksheet["B19"] = (
            f"=COUNTIF('Detailed Results'!{chunk_col}2:{chunk_col}{last_data_hint},FALSE)"
        )
        worksheet["A20"] = "Pages with DOM weight bloat"
        worksheet["B20"] = (
            f"=COUNTIF('Detailed Results'!{weight_col}2:{weight_col}{last_data_hint},\"Bloat\")"
        )

        # Delta Summary
        worksheet["A22"] = "Delta Summary"
        worksheet["A22"].fill = SECTION_FILL
        worksheet["A22"].font = Font(bold=True, color="FFFFFF")
        worksheet.merge_cells("A22:D22")

        worksheet["A23"] = "Trend status"
        worksheet["B23"] = str(delta.get("summary_line", ""))
        worksheet.merge_cells("B23:D23")
        worksheet["A24"] = "Compared against"
        worksheet["B24"] = str(delta.get("previous_name") or "(none — baseline)")
        worksheet["A25"] = "New issues"
        worksheet["B25"] = int(delta.get("new_count", 0) or 0)
        worksheet["A26"] = "Resolved issues"
        worksheet["B26"] = int(delta.get("resolved_count", 0) or 0)
        worksheet["A27"] = "Persistent issues"
        worksheet["B27"] = int(delta.get("persistent_count", 0) or 0)

        if delta.get("has_previous"):
            worksheet["B25"].fill = WARN_FILL if int(delta.get("new_count", 0) or 0) else GOOD_FILL
            worksheet["B26"].fill = GOOD_FILL
            worksheet["B27"].fill = WARN_FILL if int(delta.get("persistent_count", 0) or 0) else GOOD_FILL

        worksheet["A29"] = "New issue samples"
        worksheet["B29"] = "; ".join(list(delta.get("new_issues") or [])[:5]) or "—"
        worksheet.merge_cells("B29:D30")
        worksheet["B29"].alignment = Alignment(wrap_text=True, vertical="top")

        worksheet["A31"] = "Resolved issue samples"
        worksheet["B31"] = "; ".join(list(delta.get("resolved_issues") or [])[:5]) or "—"
        worksheet.merge_cells("B31:D32")
        worksheet["B31"].alignment = Alignment(wrap_text=True, vertical="top")

        worksheet["A34"] = "Scoring notes"
        worksheet["A34"].font = Font(bold=True)
        worksheet["A35"] = (
            "SEO, GEO, AEO, and Technical scores are per-page checklist percentages. "
            "Delta Summary compares critical failures against the latest prior workbook. "
            "Terminology checks enforce UK English and Logi-Ink brand casing. "
            "AI chunking rewards dl/dt/dd, details/summary, FAQ schema, and question headings. "
            "DOM weight flags HTML > 150 KB, > 15 scripts, or > 1,500 DOM nodes."
        )
        worksheet.merge_cells("A35:D37")
        worksheet["A35"].alignment = Alignment(wrap_text=True, vertical="top")

        worksheet.conditional_formatting.add(
            "B7:B11",
            CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=GOOD_FILL),
        )
        worksheet.conditional_formatting.add(
            "B7:B11",
            CellIsRule(operator="between", formula=["50", "79.999"], fill=WARN_FILL),
        )
        worksheet.conditional_formatting.add(
            "B7:B11",
            CellIsRule(operator="lessThan", formula=["50"], fill=BAD_FILL),
        )

        worksheet.column_dimensions["A"].width = 42
        worksheet.column_dimensions["B"].width = 36
        worksheet.column_dimensions["C"].width = 28
        worksheet.column_dimensions["D"].width = 20

        if rows:
            worksheet["C4"] = len(rows)
            worksheet["C4"].font = Font(color="6B7280", italic=True)
            worksheet["D4"] = "(static count)"
            worksheet["D4"].font = Font(color="6B7280", italic=True)

    def _build_detail(self, worksheet: Worksheet, rows: list[dict[str, Any]]) -> None:
        for col_idx, (_, header) in enumerate(DETAIL_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = THIN

        worksheet.row_dimensions[1].height = 36
        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLUMNS))}1"

        bool_cols = {
            idx
            for idx, (key, _) in enumerate(DETAIL_COLUMNS, start=1)
            if key.endswith(("_ok", "_present", "_ready", "_detected", "_complete", "_valid", "_za"))
            or key
            in {
                "redirected",
                "title_brand_ok",
                "has_organization",
                "has_local_business",
                "has_service",
                "has_website",
                "has_breadcrumb",
                "has_definition_pattern",
                "has_structured_lists",
                "has_summary_block",
                "has_same_as",
                "entity_mapping_ok",
                "nap_complete",
                "semantic_coverage_ok",
                "direct_answer_ready",
                "schema_valid",
                "heading_skip_detected",
                "local_seo_compliance",
                "json_ld_parse_fail",
                "orphan_risk",
                "link_density_ok",
                "alt_coverage_ok",
                "landmark_coverage_ok",
                "html_lang_en_za",
                "terminology_ok",
                "has_definition_list",
                "has_details_summary",
                "has_faq_schema",
                "has_question_headings",
                "ai_chunking_ready",
                "dom_weight_ok",
            }
        }

        for row_idx, data in enumerate(rows, start=2):
            for col_idx, (key, _) in enumerate(DETAIL_COLUMNS, start=1):
                value = data.get(key, "")
                if value is None:
                    value = ""
                if isinstance(value, bool):
                    value = "TRUE" if value else "FALSE"
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if col_idx in bool_cols:
                    if value == "TRUE":
                        # json_ld_parse_fail / orphan_risk TRUE is a negative signal
                        if key in {"json_ld_parse_fail", "orphan_risk"}:
                            cell.fill = BAD_FILL
                            cell.font = FALSE_FONT
                        else:
                            cell.fill = GOOD_FILL
                            cell.font = TRUE_FONT
                    elif value == "FALSE":
                        if key in {"json_ld_parse_fail", "orphan_risk"}:
                            cell.fill = GOOD_FILL
                            cell.font = TRUE_FONT
                        else:
                            cell.fill = BAD_FILL
                            cell.font = FALSE_FONT

            status_idx = self._detail_col_index("status_code")
            status_cell = worksheet.cell(row=row_idx, column=status_idx)
            try:
                code = int(data.get("status_code") or 0)
            except (TypeError, ValueError):
                code = 0
            if code >= 400 or code == 0:
                status_cell.fill = BAD_FILL
            elif 300 <= code < 400:
                status_cell.fill = WARN_FILL
            elif 200 <= code < 300:
                status_cell.fill = GOOD_FILL

            for score_key in ("seo_score", "geo_score", "aeo_score", "technical_score"):
                idx = self._detail_col_index(score_key)
                score_cell = worksheet.cell(row=row_idx, column=idx)
                try:
                    score = float(data.get(score_key) or 0)
                except (TypeError, ValueError):
                    score = 0.0
                if score >= 80:
                    score_cell.fill = GOOD_FILL
                elif score >= 50:
                    score_cell.fill = WARN_FILL
                else:
                    score_cell.fill = BAD_FILL

            title_ok_idx = self._detail_col_index("title_length_ok")
            if data.get("title_present") and not data.get("title_length_ok"):
                worksheet.cell(row=row_idx, column=title_ok_idx).fill = WARN_FILL

            heading_skip_idx = self._detail_col_index("heading_skip_detected")
            if data.get("heading_skip_detected"):
                worksheet.cell(row=row_idx, column=heading_skip_idx).fill = WARN_FILL

            # Amber highlight for South African localisation failures
            local_status_idx = self._detail_col_index("local_seo_status")
            local_cell = worksheet.cell(row=row_idx, column=local_status_idx)
            if str(data.get("local_seo_status", "")).lower() == "fail":
                local_cell.fill = WARN_FILL
                local_cell.font = Font(bold=True, color="92400E")
            elif str(data.get("local_seo_status", "")).lower() == "pass":
                local_cell.fill = GOOD_FILL
                local_cell.font = TRUE_FONT

            # Amber when missing alt text is present (accessibility gap for AI visual context)
            alt_idx = self._detail_col_index("missing_alt_count")
            try:
                alt_count = int(data.get("missing_alt_count") or 0)
            except (TypeError, ValueError):
                alt_count = 0
            if alt_count > 0:
                worksheet.cell(row=row_idx, column=alt_idx).fill = WARN_FILL

            weight_idx = self._detail_col_index("dom_weight_status")
            weight_cell = worksheet.cell(row=row_idx, column=weight_idx)
            if str(data.get("dom_weight_status", "")).lower() == "bloat":
                weight_cell.fill = WARN_FILL
            elif str(data.get("dom_weight_status", "")).lower() == "ok":
                weight_cell.fill = GOOD_FILL

        widths = {
            "url": 42,
            "title": 36,
            "og_image": 40,
            "seo_issues": 36,
            "technical_issues": 36,
            "geo_issues": 36,
            "aeo_issues": 36,
            "detected_schema_types": 32,
            "semantic_tags": 24,
            "localisation_gaps": 40,
            "local_seo_status": 28,
            "html_lang": 12,
            "terminology_issues": 36,
            "dom_weight_status": 16,
        }
        for col_idx, (key, _) in enumerate(DETAIL_COLUMNS, start=1):
            letter = get_column_letter(col_idx)
            worksheet.column_dimensions[letter].width = widths.get(key, 14)

        if rows:
            last_row = 1 + len(rows)
            for col_idx in bool_cols:
                letter = get_column_letter(col_idx)
                rng = f"{letter}2:{letter}{last_row}"
                worksheet.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f'{letter}2="FALSE"'], fill=BAD_FILL),
                )
                worksheet.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f'{letter}2="TRUE"'], fill=GOOD_FILL),
                )

            local_letter = get_column_letter(self._detail_col_index("local_seo_status"))
            local_rng = f"{local_letter}2:{local_letter}{last_row}"
            worksheet.conditional_formatting.add(
                local_rng,
                FormulaRule(formula=[f'{local_letter}2="Fail"'], fill=WARN_FILL),
            )
            worksheet.conditional_formatting.add(
                local_rng,
                FormulaRule(formula=[f'{local_letter}2="Pass"'], fill=GOOD_FILL),
            )

    @staticmethod
    def _detail_col_index(key: str) -> int:
        for idx, (column_key, _) in enumerate(DETAIL_COLUMNS, start=1):
            if column_key == key:
                return idx
        raise KeyError(key)
